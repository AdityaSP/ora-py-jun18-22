import openpyxl

wb = openpyxl.load_workbook('data.xlsx')
print wb.sheetnames
sh = wb['Sheet1']
for row in sh.columns:
    for cell in row:
        print cell.value
for col in sh.rows:
    for cell in col:
        print cell.value
wb.close()

wb = openpyxl.Workbook()
#sh = wb.create_sheet('My Data')
sh =wb[wb.sheetnames[0]]
sh['A1'] = 'hello'
sh['B1'] = 'hi'
wb.save('created.xlsx')